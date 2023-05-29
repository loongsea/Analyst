txt_A = '''
### 说明:
*  分析七八九三个年级的成绩.
*  自动生成分析报表.
'''
txt_B = '''
### 版本更新:
* ###### 2022.10.1,开始学习numpy、pandas。
* ###### 2023.2.1,完成程序设计，必须为9个学科。
* ###### 2023.4.5,用streamlit制作了一个web外壳。
* ###### 2023.4.30,重写代码，自适应若干个学科，不需要必须9个学科了。
* ###### 2023.5.7,添加滑块,确定学科分析的每班参评人数.
* ###### 2023.5.10,对学科名进行排序.
* ###### 2023.5.18,对分析结果调用模板.班级分析全班参与.
* ###### 2023.5.20,自动调用级段模板填充数据.班级分析人数可选.
'''

long_text = '''
# 设置所有学科的名称列表。
lst_xk_dic = {"语文": 1, "数学": 2, "英语": 3, "物理": 4, "化学": 5, "生物": 6, "政治": 7, "历史": 8, "地理": 9}

# 设置双达标函数的学科排名与总分排名
R_XK, R_ZF = 200, 200  # 学科排名,总分排名
# 设置双达标的积分值
JF_SDB = [7.5]

# 设置学科分数段.
ls1 = [36, 48, 60, 72, 78, 84, 90, 96, 102, 108, 114, 120]  # 120分数段
ls7 = [21, 28, 35, 42, 45, 49, 52, 56, 60, 63, 60, 70]      # 70分数段
ls5 = [15, 20, 25, 30, 32, 35, 37, 40, 42, 45, 48, 50]      # 50分数段
# 设置学科分数段的积分值.
JF_ls = [1, 2, 3, 5, 5.5, 6, 6.5, 7.5, 8, 9, 9.5, 10]       # 每个分数段的积分值

# 设置班级名次段
MC_ls = [0, 10, 50, 100, 150, 200, 250, 300, 350, 400]      # 各个名次段
# 设置班级名次段的积分值
JF_b = [10, 9.5, 9, 8.5, 8, 7, 6, 2, 2, 0]                  # 各个名次段的积分

# 成绩报表的xlsx文件.
MB_file = "c:/loongsea/模板2023.xlsx"

'''

from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import functools
import time
import os

# 设置网页信息
st.set_page_config(page_title="成绩分析_九年级_LOONGSEA", page_icon=":bar_chart:",
                   initial_sidebar_state="expanded",)
# st.title("成绩分析程序")
st.markdown(''' ## 成绩分析程序 ''')
st.write("***")
# 添加侧边栏说明文本
st.sidebar.write(txt_A)

# //////////////////////////////设置参数信息////////////////////////////////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# 设置所有学科的名称列表。
lst_xk_dic = {"语文": 1, "数学": 2, "英语": 3, "物理": 4, "化学": 5, "生物": 6, "政治": 7, "历史": 8, "地理": 9}

# 设置双达标函数的学科排名与总分排名
R_XK, R_ZF = 200, 200  # 学科排名,总分排名
# 设置双达标的积分值
JF_SDB = [7.5]

# 设置学科分数段.
ls1 = [36, 48, 60, 72, 78, 84, 90, 96, 102, 108, 114, 120]  # 120分数段
ls7 = [21, 28, 35, 42, 45, 49, 52, 56, 60, 63, 60, 70]      # 70分数段
ls5 = [15, 20, 25, 30, 32, 35, 37, 40, 42, 45, 48, 50]      # 50分数段
# 设置学科分数段的积分值.
JF_ls = [1, 2, 3, 5, 5.5, 6, 6.5, 7.5, 8, 9, 9.5, 10]       # 每个分数段的积分值

# 设置班级名次段
MC_ls = [0, 10, 50, 100, 150, 200, 250, 300, 350, 400]      # 各个名次段
# 设置班级名次段的积分值
JF_b = [10, 9.5, 9, 8.5, 8, 7, 6, 2, 2, 0]                  # 各个名次段的积分

# 成绩报表的xlsx文件.
MB_file = "c:/loongsea/模板2023.xlsx"
# //////////////////////////////设置参数信息////////////////////////////////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

# 创建一个上传文件的按钮
uploaded_file = st.file_uploader("上传XLSX文件", type=["xlsx"])

# 判断上传按钮，准备计算。
if not uploaded_file:
    st.text_area(label='设置说明', value=long_text, height=570, help=txt_B)
elif uploaded_file:
    start = time.time()
    # ////////////////////////////////////////////读取excel文件，存储为df表//////////////////////////////////////////////////////
    # /////////////////////////////////////////////添加主要界面控件//////////////////////////////////////////////////////////////

    # 添加两个分列.
    col_A, col_B = st.columns(2)

    # 左分列
    with col_A:
        # 使用Pandas读取己上传的Excel文件
        df = pd.read_excel(uploaded_file, engine='openpyxl', sheet_name=None)      # 读取所有工作表。
        # 列出所有工作表的名称，
        sheet_names = list(df.keys())
        # 创建一个选择框，返回列表中的一个选中的工作表。
        selected_sheet = st.selectbox("选择工作表", sheet_names)
        # 将所选工作表的数据返回df表。
        df = df[selected_sheet]                  # *****************df表为所选工作表学生成绩表***********************
        # 添加一个滑动条,用于选择统计学科成绩时,计算的班级学生数.
        bc = st.slider('选择学科分析范围', min_value=1, max_value=60, value=45)

    # 右分列:
    with col_B:
        # 读取模板文件
        df_MB = pd.read_excel(MB_file, engine='openpyxl', sheet_name=None)
        # 列出模板工作表的名称。
        sht_MB_names = list(df_MB.keys())
        # 创建选择框，返回其中选中的工作表。
        sht_MB_names = st.selectbox("选择报表模板", sht_MB_names)
        # 依据名称，返回模板工作表
        df_MB = df_MB[sht_MB_names]             # ***************df_MB表为模板文件所对应年级的工作表******************
        # 添加一个滑动条,用于选择统计学科成绩时,计算的班级学生数.
        bc_all = st.slider('选择班级分析范围', min_value=1, max_value=60, value=50)

    # # 添加一个滑动条,用于选择班次.
    # bc = st.slider('选择学科统计范围', min_value=1, max_value=60, value=45)

    # ///////////////////////////////////////////创建df表，df_MC_XK表，df_MC_XKZF表，////////////////////////////////////////////
    # ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

    # df表列标的列表
    lst_df_col = df.columns
    # 确定学科列标的列表，即为学科名称列表与df表列标列表的交集。
    lst_xk_all = lst_xk_dic.keys()
    lst_XK = list(set(lst_xk_all) & set(lst_df_col))
    # 对学科名进行排序
    lst_XK.sort(key=lambda x: lst_xk_dic[x])       # ****************学科列标的列表lst_XK：[语文，数学...地理]***************
    # 确定学科总分的列标列表
    lst_XKZF = lst_XK + ["总分"]                    # **********学科总分列标的列表lst_XKZF：[语文，数学...地理,总分]************

    # 为df表添加总分列,df表中将有总分列。
    df["总分"] = df.loc[:, lst_XK].sum(axis=1)      # 增加总分列
    # 添加级次列
    df["级次"] = df["总分"].rank(axis=0, ascending=False, method="min")  # 增加级次列,从大到小,同名次取最小值.
    # 添加班次列,从大到小,取最小值.
    df["班次"] = df.groupby("班级")["总分"].rank(axis=0, ascending=False, method="min")  # df表完成,包括班级,学号,姓名,各学科,总分,班级,级次.

    # 复制df表为df_all,所括了所有学生的学科分数等信息
    df_all = df.copy()
    df_all = df_all[df_all["班次"] <= bc_all]
    df_all_MC_XK = df.loc[:, lst_XKZF].rank(axis=0, method="min", ascending=False)  # 各科与总分的名次
    df_all_MC_XK = pd.concat([df_all[["班级", "学号"]], df_all_MC_XK], axis=1)       # 班级，学号，各科名次，总分名次。
    df_all_RS = df_all_MC_XK.groupby("班级").agg({"总分": 'count',"学号": 'count'})                        # 分班统计总分的个数,,也就是每班人数

    # 返回将df表中班次列的值小于bc的数据
    df = df[df["班次"] <= bc]
    # 显示工作表数据。
    st.dataframe(df, height=230)

    # 将学科与总分列的值转化为名次
    df_MC_XK = df.loc[:, lst_XKZF].rank(axis=0, method="min", ascending=False)  # 各科与总分的名次
    # 为学科名次表添加班级与学号列。
    df_MC_XK = pd.concat([df[["班级", "学号"]], df_MC_XK], axis=1)  # 学科名次表df_MC_XK完成.包括班级,学号,各学科及总分的名次.

    # 确定学科总分列表的总列数
    num_XKZF = len(lst_XKZF) - 1
    # 创建学科总分名次表df_XKZF.
    df_MC_XKZF = df.loc[:, lst_XKZF].rank(axis=0, method="min", ascending=False)  # 各科与总分的名次
    # 将每个学科的名次与总分名次合并为一个元组
    for i in range(num_XKZF):
        df_MC_XKZF.iloc[:, i] = df_MC_XKZF.iloc[:, [i, num_XKZF]].apply(tuple, axis=1)  # 将多列的名次数据与总分名次列合并为多个元组列.
    # 为学科总分名次表添加班级与学号列。
    df_MC_XKZF = pd.concat([df[["班级", "学号"]], df_MC_XKZF], axis=1)  # 学科总分名次表df_MC_XKZF完成.包括班级,学号,各学科及总分的名次组成的元组.


    # ///////////////////////////////////////////自定义双达标函数，对学科总分名次表进行双达标运算///////////////////////////////////////
    # ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # 定义双达标函数,要求学科排名小于m,总分排名n.
    def SDB(sr, m, n):
        s = 0
        for i in sr:
            if (i[0] <= m) & (i[1] <= n):
                s += 1
        return s


    # 定义偏函数，设置学科排名为R_XK,总分排名为R_ZF.
    SDB_A = functools.partial(SDB, m=R_XK, n=R_ZF)  # 设置偏函数SDB_A,m/n代表学科名次/总分名次.

    # 创建字典，键值对为{"总分":'count'},统计总分的个数,以得到参评人数。
    dic_XKZF = {"总分": 'count'}
    # 更新字典，添加键值对为{学科名：双达标函数},以计算双达标人数.
    dic_XKZF.update({str(i): SDB_A for i in lst_XK})  # {"总分": 'count', "语文": SDB_A .......})

    # 使用聚合函数,确定参评人数和双达标人数.
    df_SDB = df_MC_XKZF.groupby("班级").agg(dic_XKZF)  # 等价于.agg({"总分": 'count', "语文": SDB_A .......})


    # ///////////////////////////////////////////分段统计函数组，两率函数组////////////////////////////////////////////////////////
    # ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # 定义分段函数,函数的反回值是一个函数组,每个函数统计落在区间的分数的个数.
    def Get_Func(ls):  # 数组中落在列表数据区间中的数据个数.
        funcs = []
        for i in range(len(ls)):
            if i == len(ls) - 1:
                exec(f"def F_{i + 1}(arr,p={i}):return sum(arr >= {ls}[p])")
            else:
                exec(f"def F_{i + 1}(arr,p={i}):return sum((arr >= {ls}[p]) & (arr < {ls}[p + 1]))")
            funcs.append(locals()[f"F_{i + 1}"])
        return funcs


    # 定义偏函数,分别对应120,70,50三种学科分数的情况.ls1,ls7,ls5的值在最前的设置中修改.
    Get_Func1 = functools.partial(Get_Func, ls=ls1)  # 设置偏函数Get_Func1,预先设置Get_Func的ls值.
    Get_Func7 = functools.partial(Get_Func, ls=ls7)  # 设置偏函数Get_Func7,预先设置Get_Func的ls值.
    Get_Func5 = functools.partial(Get_Func, ls=ls5)  # 设置偏函数Get_Func5,预先设置Get_Func的ls值.
    Funcs1 = Get_Func1()  # 对120分值学科进行分数段计数的函数组
    Funcs7 = Get_Func7()  # 对70分值学科进行分数段计数的函数组
    Funcs5 = Get_Func5()  # 对50分值学科进行分数段计数的函数组

    # 120分值的两率函数，包括人数与比率。
    P72T120 = lambda arr: sum((arr >= 72) & (arr <= 120))  # 数组中72<=x<120的数据个数
    P72L120 = lambda arr: (sum((arr >= 72) & (arr <= 120))) / len(arr)  # 数组中72<=x<120的数据的占比.
    P96T120 = lambda arr: sum((arr >= 96) & (arr <= 120))  # 数组中96<=x<120的数据个数
    P96L120 = lambda arr: (sum((arr >= 96) & (arr <= 120))) / len(arr)  # 数组中96<=x<120的数据的占比.

    # 70分值的两率函数，包括人数与比率。
    P42T70 = lambda arr: sum((arr >= 42) & (arr <= 70))  # 数组中及格的数据个数
    P42L70 = lambda arr: (sum((arr >= 42) & (arr <= 70))) / len(arr)  # 数组中及格的数据的占比.
    P56T70 = lambda arr: sum((arr >= 56) & (arr <= 70))  # 数组中优秀的数据个数
    P56L70 = lambda arr: (sum((arr >= 56) & (arr <= 70))) / len(arr)  # 数组中优秀的数据的占比.

    # 50分值的两率函数，包括人数与比率。
    P30T50 = lambda arr: sum((arr >= 30) & (arr <= 50))  # 数组中72<=x<120的数据个数
    P30L50 = lambda arr: (sum((arr >= 30) & (arr <= 50))) / len(arr)  # 数组中72<=x<120的数据的占比.
    P40T50 = lambda arr: sum((arr >= 40) & (arr <= 50))  # 数组中96<=x<120的数据个数
    P40L50 = lambda arr: (sum((arr >= 40) & (arr <= 50))) / len(arr)  # 数组中96<=x<120的数据的占比.

    # 分数段函数与两率一平函数组。
    cont120 = [*Funcs1, P72T120, P72L120, P96T120, P96L120, "mean"]
    cont70 = [*Funcs7, P42T70, P42L70, P56T70, P56L70, "mean"]
    cont50 = [*Funcs5, P30T50, P30L50, P40T50, P40L50, "mean"]


    # ///////////////////////////////////////////应用分段统计函数与两率一平函数，生成分数段表//////////////////////////////////////////
    # ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # 根据不同的学科，应用不同的分段函数与二率一平方案。
    def Get_FSD(sr):
        if sr in ["语文", "数学", "英语"]:  # ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]
            return cont120
        elif sr in ["物理", "政治"]:
            return cont70
        elif sr in ["化学", "生物", "历史", "地理"]:
            return cont50
        else:
            pass


    # 定义一个分数段表df_FSD，包括班级，和各个学科的分值。
    df_FSD = pd.concat([df["班级"], df.loc[:, lst_XK]], axis=1)  # 只包括班级列/各学科列
    # 创建字典，键值对为"语文":cont120,。
    dic_FSD = {str(i): Get_FSD(i) for i in lst_XK}  #
    # 按班级分组后，执行分数组函数与两率一平计算。
    df_FSD = df_FSD.groupby("班级").agg(dic_FSD)

    # ///////////////////////////////////////////生成各科报表///////////////////////////////////////////////////////////////////
    # ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # 确定双达标与分数段的积分值,由双达标的权得值与分数值的权重值列表组合而成.
    JF_ls_SDB = JF_SDB + JF_ls  # 双达标积分表+分数段的积分表


    # 定义函数,对分数段的人数与权得值进行累计积分的计算.
    def JF(arr, qz):  # 计算总积分的函数
        res = 0
        for m, n in zip(arr, qz):
            res += m * n
        return res


    # 为函数设置偏好值,将权重值先行加入.
    JF = functools.partial(JF, qz=JF_ls_SDB)  # 设置偏函数,预先设置qz的值为qzjf.

    # 按学科列表生成学科报表,并加入积分列与排名列.
    for xk in lst_XK:
        B_name = f"B_{xk}"
        locals()[B_name] = pd.concat([df_SDB[["总分", xk]], df_FSD[xk]], axis=1)  # 生成报表,xk实际上是双达标表中的双达标的数据.
        locals()[B_name].insert(loc=1, column="积分",
                                value=locals()[B_name].iloc[:, 1:(len(JF_ls_SDB) + 1)].apply(JF, axis=1))  # 添加积分列
        locals()[B_name].insert(loc=2, column="排名",
                                value=locals()[B_name]["积分"].rank(axis=0, ascending=False, method="min"))  # 添加积分排名列
        # 对班级表的部分列标得命名。
        locals()[B_name].rename(columns={"总分": "人数", "<lambda_0>": "及格人数", "<lambda_1>": "及格率",
                                         "<lambda_2>": "优率人数", "<lambda_3>": "优秀率", "mean": "平均分"}, inplace=True)


    # ///////////////////////////////////////////计算并生成班级报表///////////////////////////////////////////////////////////////
    # ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # 定义名次段函数。
    def MC_Func(ls):  # 数组中落在列表数据区间中的数据个数.
        funcs = []
        for i in range(len(ls)):
            if i == len(ls) - 1:
                exec(f"def M_{i + 1}(arr,p={i}):return sum(arr > {ls}[p])")
            else:
                exec(f"def M_{i + 1}(arr,p={i}):return sum((arr > {ls}[p]) & (arr <= {ls}[p + 1]))")
            funcs.append(locals()[f"M_{i + 1}"])
        return funcs

    # 为名次段函数设置偏函数，缺省值为名次表。
    MC_Func1 = functools.partial(MC_Func, ls=MC_ls)  # 设置偏函数Get_Func1,预先设置Get_Func的ls值.
    # 生成名次段函数组。
    MC_FuncsA = MC_Func1()  # 生成名次在0-400之间个数统计函数组.

    # 定义函数，统计前10名，前50名，前100名等的人数。
    Q10 = lambda arr: sum(arr <= 10)
    Q50 = lambda arr: sum(arr <= 50)
    Q100 = lambda arr: sum(arr <= 100)
    Q150 = lambda arr: sum(arr <= 150)
    Q200 = lambda arr: sum(arr <= 200)
    Q250 = lambda arr: sum(arr <= 250)
    Q300 = lambda arr: sum(arr <= 300)

    # 为积分函数设置偏好值，设置为班级积分表
    BJF = functools.partial(JF, qz=JF_b)  # 设置偏函数,预先设置qz的值为qzjf.

    # 创建班级表，只包括班级列与总分列，班级用来分组，总分用来应用名次段函数。
    df_BJ = df_all_MC_XK [["班级", "总分"]]  # 将df_MC复制为df_BJ.    df_all_MC_XK
    # 用班级分组，用总分应用名次段函数。
    df_BJ = df_BJ.groupby("班级")["总分"].agg([*MC_FuncsA, Q10, Q50, Q100, Q150, Q200, Q250, Q300])
    # 插入积分列。
    df_BJ.insert(loc=0, column="积分", value=df_BJ.iloc[:, 0:len(MC_ls)].apply(BJF, axis=1))  # 添加积分列
    # 插入名次列。
    df_BJ.insert(loc=1, column="排名", value=df_BJ["积分"].rank(axis=0, ascending=False, method="min"))  # 添加积分排名列
    # 插入双达标的总分，其实是总人数。
    # df_BJ = pd.concat([df_SDB["总分"], df_BJ], axis=1)    # df_all_RS
    df_BJ = pd.concat([df_all_RS["总分"], df_BJ], axis=1)  # df_all_RS      在这儿.
    # 对班级表的部分列标得命名。
    df_BJ.rename(
        columns={"总分": "人数", "<lambda_0>": "前10", "<lambda_1>": "前50", "<lambda_2>": "前100", "<lambda_3>": "前150",
                 "<lambda_4>": "前200", "<lambda_5>": "前250", "<lambda_6>": "前300"}, inplace=True)

    # ///////////////////////////////////////将df表及其它表存储为excel文件，并装入下载按钮////////////////////////////////////////////
    # //////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

    # 打开模板文件。
    wb = load_workbook(MB_file)  # 模板文件.
    for sht in wb.worksheets:  # 删除非选定文件.
        if sht.title != sht_MB_names:
            wb.remove(sht)
    ws = wb[sht_MB_names]  # 声明模板文件中的工作表.


    # 定义函数:把pd表的值注入ws工作表的指定行列。
    def df_to_ws(ws, row, col, df):
        '''
        ws:目标工作表(openpyxl的工作表)，
        row:目标工作表的行，
        col:目标工作表的列，
        df:dataframe数据表，由pandas生成。
        '''
        # 确定df表的行数，列数。
        HS = df.shape[0]
        LS = df.shape[1]
        # 将df表转化为列表。
        df_arr = np.array(df)
        df_lst = df_arr.tolist()
        # 把df_lst注入ws表。
        for r in range(HS):
            for c in range(LS):
                ws.cell(r + row, c + col, df_lst[r][c])

    # 判断是哪个年级。
    if ("九年级" in sht_MB_names) or ("七年级" in sht_MB_names) :
        loc_BJ = 117
    elif ("八年级" in sht_MB_names):
        loc_BJ = 133
    else:
        loc_BJ = 133
    # 将年级的班级统计数据注入ws表中。
    df_to_ws(ws, loc_BJ, 4, df_BJ)

    # 遍历学科报表,并将学科报表注入wb表：
    JG = 0  # 学科表的行号.
    for xk in lst_XK:
        # 获取学科df表名称.
        B_name = f"B_{xk}"
        # 获取学科df表对象.
        df_XK = locals()[B_name]  # df_XK为学科报表。
        # 把学科的df表数据注入ws表中.
        df_to_ws(ws, 5 + JG, 4, df_XK)
        JG += 16

    # 创建一个BytesIO对象excel_file,用来存储Excel文件
    excel_file = BytesIO()
    # 保存新工作簿到IO文件中.
    wb.save(excel_file)

    # ///////////////////////////////////////添加下载按钮,以下载BytesIO对象中己注入ws的数据文件(excel_file)///////////////////////////
    # //////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # 添加一个运算成功信息框。
    end = time.time()
    tim = end - start
    st.success(f"运算己经完成，共用时：{round(end - start, 2)}秒。")

    # 创建下载按钮
    st.download_button(
        label='下载分析结果',
        data=excel_file,
        file_name=os.path.splitext(uploaded_file.name)[0] + "_" + sht_MB_names + "_报表" + ".xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


